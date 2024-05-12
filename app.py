from GPT import process_prompt
import chainlit as cl
from langchain.document_loaders import Docx2txtLoader
import PyPDF2
from io import BytesIO
from chainlit.action import Action
from chainlit.input_widget import Select, Switch, Slider
import openpyxl

is_magic_prompt = False
author_name = "BOT"

@cl.on_chat_start
async def start_chat():
    await cl.Avatar(
        name=author_name,
        url="./public/bot.png",
    ).send()

    await cl.Avatar(
        name="User",
        url="./public/human.png",
    ).send()

    cl.user_session.set(
        "message_history_obj",
        [{"role": "system", "content": "You are a helpful assistant. You need to answer questions based on the input_data: provided. If the context of the question is outside of the input_data: then respond politely respond that you are not aware of the answer."}],
    )
    files = None
    while files is None:
        files = await cl.AskFileMessage(
            content="Please upload a .docx, .pdf or .xlsx file to begin!",
            accept=["application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/pdf", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
            max_size_mb=5,
            timeout=180,
            author=author_name,
            max_files=2
        ).send()

    doc_text = ""
    for index, file in enumerate(files):
    
        msg = cl.Message(content=f"Processing `{file.name}`...", author=author_name)
        await msg.send()

        if file.name.split(".")[-1] == "pdf":
            pdf_stream = BytesIO(file.content)
            pdf = PyPDF2.PdfReader(pdf_stream)
            doc_text += f" file_name: {file.name}, input_data: "
            for page in pdf.pages:
                doc_text += page.extract_text()

        elif  file.name.split(".")[-1] == "docx":
            with open("saved_file.docx", "wb") as f:
                f.write(file.content)
            f.close()

            loader = Docx2txtLoader("saved_file.docx")
            documents = loader.load()
            doc_text += f" file_name: {file.name}, input_data: "
            for document in documents:
                doc_text += document.page_content

        elif  file.name.split(".")[-1] == "xlsx":
            with open("saved_file.xlsx", "wb") as f:
                f.write(file.content)
            f.close()
            
            workbook = openpyxl.load_workbook("saved_file.xlsx")
           
            plain_text_data = ""
            doc_text += f" file_name: {file.name}, "
            for sheet_name in workbook.sheetnames:
                doc_text += f"worksheet_name: {sheet_name}, input_data: "
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(values_only=True):
                    for index, cell_value in enumerate(row):
                        plain_text_data += str(cell_value)
                        
                        if index < len(row) - 1:
                            plain_text_data += ", "
                    plain_text_data += " \n "

            doc_text += plain_text_data

    cl.user_session.set(
        "message_history_obj",
        [{"role": "system", "content": doc_text}],
    )
    
    print(doc_text)

    file_names = ""
    for file in files:
        if len(file_names) != 0:
            file_names += " and "
        file_names += file.name
    await cl.Message(content=f"Analysing `{file_names}` is completed. You can now ask questions!", author=author_name).send()
    
    

@cl.on_message
async def main(message: str):
    message_history_obj = cl.user_session.get("message_history_obj")
    message_history_obj.append({"role": "user", "content": message})
    response = process_prompt(message_history_obj)
    message_history_obj.append({"role": "assistant", "content": response})
    await cl.Message(
        content=f"{response}", author=author_name
    ).send()
